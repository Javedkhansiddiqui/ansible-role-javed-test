# importing openpyxl module
import openpyxl

# workbook object is created
wb_obj = openpyxl.load_workbook(r"template/read_excel_file.xlsx")

sheet_obj = wb_obj.active

data = sheet_obj.cell(row=2, column=2)

print(data.value)

# max_col = sheet_obj.max_column

# # Will print a particular row value
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=2, column=i)
#     print(cell_obj.value, end=" ")
